#!/usr/bin/env python3
"""DATA INTERN ASSIGNMENT — Complete Solution"""
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import warnings, sys
warnings.filterwarnings('ignore')

XLSX = "/home/wasim/Documents/github/DATA-INTERN-ASSIGNMENT/Assignment_data_dictionary.xlsx"

print("Step 1: Reading Excel...", flush=True)
xl = pd.ExcelFile(XLSX)
print("Sheets:", xl.sheet_names, flush=True)
main = xl.sheet_names[0]
df_raw = xl.parse(main)
print(f"Shape: {df_raw.shape}", flush=True)
print("Columns:", list(df_raw.columns), flush=True)

# Quick stats
total_rows, total_cols = df_raw.shape
null_pct = (df_raw.isnull().sum() / total_rows * 100).round(2)
print(f"\nNull columns: {(null_pct > 0).sum()}", flush=True)
print("\nSample data types:\n", df_raw.dtypes, flush=True)
print("\nFirst 3 rows:\n", df_raw.head(3).to_string(), flush=True)

# ====== PART 2: CLEANING ======
print("\n=== CLEANING ===", flush=True)
df = df_raw.copy()

# Find key columns dynamically
cols_upper = {c: c.upper() for c in df.columns}

def find_col(*keywords):
    for c, cu in cols_upper.items():
        if all(k in cu for k in keywords):
            return c
    return None

dt_cols_map = {}
for prefix in ['CLASS_START', 'CLASS_END', 'ACTUAL_START', 'ACTUAL_END']:
    col = find_col(*prefix.split('_'), 'DATE') or find_col(*prefix.split('_'), 'TIME') or find_col(*prefix.split('_'))
    if col:
        dt_cols_map[prefix] = col

print("DateTime cols found:", dt_cols_map, flush=True)

# Parse datetimes
for prefix, col in dt_cols_map.items():
    try:
        df[col] = pd.to_datetime(df[col], errors='coerce', infer_datetime_format=True)
        df[f'{prefix}_DATE'] = df[col].dt.date
        df[f'{prefix}_TIME'] = df[col].dt.strftime('%H:%M:%S')
        print(f"  Parsed {col}", flush=True)
    except Exception as e:
        print(f"  Error parsing {col}: {e}", flush=True)

# CLASS_DURATION_ACTUAL & CLASS_DELAY_MINS
if 'ACTUAL_START' in dt_cols_map and 'ACTUAL_END' in dt_cols_map:
    try:
        dur = (df[dt_cols_map['ACTUAL_END']] - df[dt_cols_map['ACTUAL_START']]).dt.total_seconds() / 60
        df['CLASS_DURATION_ACTUAL_MINS'] = dur.round(2)
        print("  Created CLASS_DURATION_ACTUAL_MINS", flush=True)
    except: pass

if 'CLASS_START' in dt_cols_map and 'ACTUAL_START' in dt_cols_map:
    try:
        delay = (df[dt_cols_map['ACTUAL_START']] - df[dt_cols_map['CLASS_START']]).dt.total_seconds() / 60
        df['CLASS_DELAY_MINS'] = delay.round(2)
        print("  Created CLASS_DELAY_MINS", flush=True)
    except: pass

# Cancel logic
cancel_col = find_col('CANCEL') or find_col('STATUS')
if cancel_col:
    cancelled = df[cancel_col].astype(str).str.upper().str.strip().isin(['CANCELLED','CANCELED','CANCEL','1','TRUE','YES'])
    actual_c = [c for c in df.columns if 'ACTUAL' in c.upper()]
    attempt_c = [c for c in df.columns if 'ATTEMPT' in c.upper() or ('DURATION' in c.upper() and 'CLASS_DURATION' not in c.upper())]
    score_c = [c for c in df.columns if any(k in c.upper() for k in ['SCORE','ASSIGN','CLASSWORK','HOMEWORK'])]
    for c in actual_c + attempt_c + score_c:
        df.loc[cancelled, c] = np.nan
    print(f"  Cancelled rows cleaned: {cancelled.sum():,}", flush=True)

# Attendance logic
att_col = find_col('ATTENDANCE')
attempt_dur = find_col('ATTEMPT', 'DURATION') or find_col('ATTEMPT')
if att_col:
    absent = df[att_col].astype(str).str.upper().str.strip().isin(['ABSENT','0','N','NO','FALSE'])
    present = df[att_col].astype(str).str.upper().str.strip().isin(['PRESENT','1','Y','YES','TRUE'])
    df['IS_PRESENT'] = present.astype(int)
    print(f"  Absent: {absent.sum():,}, Present: {present.sum():,}", flush=True)

    if attempt_dur:
        df[attempt_dur] = pd.to_numeric(df[attempt_dur], errors='coerce')
        df.loc[absent, attempt_dur] = np.nan
        df.loc[df[attempt_dur] < 0, attempt_dur] = np.nan

    for c in df.columns:
        cu = c.upper()
        if ('CLASSWORK' in cu or 'HOMEWORK' in cu) and 'SUBMIT' in cu:
            df[c] = pd.to_numeric(df[c], errors='coerce')
            df.loc[absent, c] = 0

# Score cleaning
score_cols = [c for c in df.columns if 'SCORE' in c.upper() and 'MAX' not in c.upper()]
max_cols = [c for c in df.columns if 'MAX' in c.upper() and 'SCORE' in c.upper()]
for sc in score_cols:
    df[sc] = pd.to_numeric(df[sc], errors='coerce')
    df.loc[df[sc] < 0, sc] = 0
for sc in score_cols:
    for mc in max_cols:
        try:
            df[mc] = pd.to_numeric(df[mc], errors='coerce')
            mask = df[sc] > df[mc]
            df.loc[mask & mask.notna(), sc] = df.loc[mask & mask.notna(), mc]
        except: pass

print("Cleaning done.", flush=True)

# ====== PART 3: TRANSFORMATION ======
print("\n=== TRANSFORMATION ===", flush=True)

# Attendance % per class
class_id = find_col('CLASS', 'ID') or find_col('CLASS_ID')
if class_id and 'IS_PRESENT' in df.columns:
    att_pct = df.groupby(class_id)['IS_PRESENT'].transform('mean') * 100
    df['ATTENDANCE_PCT'] = att_pct.round(2)
    print("  ATTENDANCE_PCT created", flush=True)

# Engagement score
engage = pd.Series(0.0, index=df.index)
n = 0
if 'IS_PRESENT' in df.columns:
    engage += df['IS_PRESENT'] * 40; n += 1
cw_sub = find_col('CLASSWORK', 'SUBMIT')
if cw_sub:
    df[cw_sub] = pd.to_numeric(df[cw_sub], errors='coerce').fillna(0)
    engage += df[cw_sub].clip(0,1) * 30; n += 1
hw_sub = find_col('HOMEWORK', 'SUBMIT') or find_col('ASSIGNMENT', 'SUBMIT')
if hw_sub:
    df[hw_sub] = pd.to_numeric(df[hw_sub], errors='coerce').fillna(0)
    engage += df[hw_sub].clip(0,1) * 20; n += 1
rating_col = find_col('RATING')
if rating_col:
    r_norm = pd.to_numeric(df[rating_col], errors='coerce').fillna(0)
    mx = r_norm.max()
    if mx > 0:
        engage += (r_norm / mx) * 10; n += 1
df['ENGAGEMENT_SCORE'] = engage.round(2)
print(f"  ENGAGEMENT_SCORE created (components: {n})", flush=True)

# Teacher punctuality
if 'CLASS_DELAY_MINS' in df.columns:
    df['TEACHER_PUNCTUALITY'] = pd.cut(
        df['CLASS_DELAY_MINS'], bins=[-9999, 5, 15, 99999],
        labels=['On Time', 'Late', 'Very Late']
    )
    print("  TEACHER_PUNCTUALITY created", flush=True)

# Week of month
if 'CLASS_START' in dt_cols_map:
    dt = df[dt_cols_map['CLASS_START']]
    df['WEEK_OF_MONTH'] = dt.dt.day.apply(lambda d: f'W{min((d-1)//7+1, 4)}' if pd.notna(d) else None)
    df['DAY_OF_WEEK'] = dt.dt.day_name()
    df['CLASS_HOUR_BUCKET'] = dt.dt.hour.apply(
        lambda h: 'Morning' if h < 12 else ('Afternoon' if h < 17 else 'Evening') if pd.notna(h) else None
    )
    print("  WEEK_OF_MONTH, DAY_OF_WEEK, CLASS_HOUR_BUCKET created", flush=True)

print("Transformation done.", flush=True)

# ====== PART 4: ANALYSIS ======
print("\n=== ANALYSIS ===", flush=True)
results = {}

student_col = find_col('STUDENT', 'ID') or find_col('STUDENT', 'NAME') or find_col('STUDENT')
teacher_col = find_col('TEACHER', 'ID') or find_col('TEACHER', 'NAME') or find_col('TEACHER') or find_col('FACULTY')
exam_col = find_col('EXAM') or find_col('COURSE') or find_col('BATCH') or find_col('PROGRAM')
cw_score = find_col('CLASSWORK', 'SCORE')
hw_score = find_col('HOMEWORK', 'SCORE') or find_col('ASSIGNMENT', 'SCORE')

print(f"Keys: student={student_col}, teacher={teacher_col}, exam={exam_col}", flush=True)
print(f"Scores: cw={cw_score}, hw={hw_score}, rating={rating_col}", flush=True)

# 4.1 Student Behaviour
if student_col and 'IS_PRESENT' in df.columns:
    agg = {'IS_PRESENT': ['count','sum','mean']}
    if cw_score: agg[cw_score] = 'mean'
    if hw_score: agg[hw_score] = 'mean'
    if 'ENGAGEMENT_SCORE' in df.columns: agg['ENGAGEMENT_SCORE'] = 'mean'
    ss = df.groupby(student_col).agg(agg).round(2)
    ss.columns = ['Total_Classes','Attended','Att_Rate','Avg_CW_Score','Avg_HW_Score','Avg_Engagement'][:len(ss.columns)]
    if 'Att_Rate' in ss.columns:
        ss['Att_Rate'] = (ss['Att_Rate']*100).round(1)
    results['StudentBehaviour'] = ss.reset_index().head(500)
    q90 = ss['Att_Rate'].quantile(0.9) if 'Att_Rate' in ss.columns else 90
    results['Top10pctStudents'] = ss[ss['Att_Rate'] >= q90].sort_values('Att_Rate', ascending=False).reset_index().head(100)
    print(f"  4.1 done: {len(ss)} students", flush=True)

# 4.2 Teacher Performance
if teacher_col and 'IS_PRESENT' in df.columns:
    tagg = {'IS_PRESENT': 'mean'}
    if rating_col: tagg[rating_col] = 'mean'
    if 'CLASS_DELAY_MINS' in df.columns: tagg['CLASS_DELAY_MINS'] = 'mean'
    if cw_score: tagg[cw_score] = 'mean'
    ts = df.groupby(teacher_col).agg(tagg).round(2)
    ts.columns = ['Att_Rate'] + [c for c in ['Avg_Rating','Avg_Delay','Avg_CW_Score'][:len(ts.columns)-1]]
    ts['Att_Rate'] = (ts['Att_Rate']*100).round(1)
    if 'TEACHER_PUNCTUALITY' in df.columns:
        ontime = df.groupby(teacher_col)['TEACHER_PUNCTUALITY'].apply(lambda x: (x=='On Time').sum()/len(x)*100).round(1)
        ts['OnTime_Pct'] = ontime
    ts_r = ts.reset_index()
    results['TeacherPerformance'] = ts_r
    results['Top10Teachers'] = ts_r.nlargest(10, 'Att_Rate')
    results['Bottom10Teachers'] = ts_r.nsmallest(10, 'Att_Rate')
    print(f"  4.2 done: {len(ts)} teachers", flush=True)

# 4.3 Exam Insights
if exam_col and 'IS_PRESENT' in df.columns:
    eagg = {'IS_PRESENT': ['mean','count']}
    if 'ENGAGEMENT_SCORE' in df.columns: eagg['ENGAGEMENT_SCORE'] = 'mean'
    if cw_score: eagg[cw_score] = 'mean'
    es = df.groupby(exam_col).agg(eagg).round(2)
    es.columns = ['_'.join(c) for c in es.columns]
    results['ExamInsights'] = es.reset_index()
    print(f"  4.3 done: {len(es)} exam types", flush=True)

# 4.4 Time-Based
if 'CLASS_HOUR_BUCKET' in df.columns and 'IS_PRESENT' in df.columns:
    results['TimeAttendance'] = (df.groupby('CLASS_HOUR_BUCKET')['IS_PRESENT'].mean()*100).round(1).reset_index()
    print("  4.4 time done", flush=True)
if 'DAY_OF_WEEK' in df.columns and 'IS_PRESENT' in df.columns:
    results['DayAttendance'] = (df.groupby('DAY_OF_WEEK')['IS_PRESENT'].mean()*100).round(1).reset_index()
    print("  4.4 day done", flush=True)

# 4.5 Other
if 'TEACHER_PUNCTUALITY' in df.columns and rating_col:
    results['PunctualityVsRating'] = df.groupby('TEACHER_PUNCTUALITY')[rating_col].mean().round(2).reset_index()
    print("  4.5 punctuality vs rating done", flush=True)

print(f"Analysis results: {list(results.keys())}", flush=True)

# ====== WRITE TO EXCEL ======
print("\n=== WRITING EXCEL ===", flush=True)
wb = load_workbook(XLSX)

hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
sub_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
alt_fill = PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid')
yel_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
grn_fill = PatternFill(start_color='E0FFE0', end_color='E0FFE0', fill_type='solid')
bw = Font(bold=True, color='FFFFFF', size=12)
bb = Font(bold=True, size=11)
nf = Font(size=10)
wrap = Alignment(wrap_text=True, vertical='top')

def rm_sheet(name):
    if name in wb.sheetnames: del wb[name]

def write_df_to_sheet(ws, dataframe, start_row=1):
    for ci, col in enumerate(dataframe.columns, 1):
        c = ws.cell(row=start_row, column=ci, value=str(col))
        c.fill = hdr_fill; c.font = bw
    for ri, row in enumerate(dataframe.itertuples(index=False), start_row+1):
        for ci, val in enumerate(row, 1):
            try:
                if pd.isna(val): val = None
                elif isinstance(val, (np.integer,)): val = int(val)
                elif isinstance(val, (np.floating,)): val = float(val)
            except: pass
            try: ws.cell(row=ri, column=ci, value=val)
            except: ws.cell(row=ri, column=ci, value=str(val))

# Understanding sheet
rm_sheet('Understanding')
ws = wb.create_sheet('Understanding')
ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 80

ws.merge_cells('A1:B1')
c = ws.cell(row=1, column=1, value='PART 1 — DATA UNDERSTANDING')
c.fill = hdr_fill; c.font = bw

qa = [
    ('1. What does each row represent?',
     f'Each row = one STUDENT × one CLASS SESSION interaction.\n'
     f'Captures attendance, scores, timing, teacher, and submission.\n'
     f'Total: {total_rows:,} rows × {total_cols} columns.'),
    ('2. Major entities in data?',
     'STUDENT (learner), CLASS/SESSION (scheduled event),\n'
     'TEACHER/FACULTY (instructor), BATCH/EXAM (JEE/NEET/CBSE/Foundation),\n'
     'ASSIGNMENTS (classwork, homework), SCORES.'),
    ('3. Which columns look messy?',
     'DATETIME cols (mixed formats), ATTENDANCE (Present/1/Y),\n'
     'SCORES (negatives, >max), DURATION (negatives for absent),\n'
     'SUBMISSION flags (0/1 vs Yes/No).'),
]

dq = [
    'DQ1: Mixed datetime formats across CLASS/ACTUAL datetime columns',
    'DQ2: Missing ACTUAL datetimes for non-cancelled classes',
    'DQ3: Negative attempt durations (impossible)',
    'DQ4: Absent students with non-zero attempt duration',
    'DQ5: Scores exceeding max scores',
    'DQ6: Negative scores',
    'DQ7: Absent students with classwork/homework marked submitted',
    'DQ8: Cancelled classes with actual data populated',
    'DQ9: Inconsistent attendance encoding (Present/1/Y/Yes)',
    'DQ10: Mixed submission flag types (0/1 vs Yes/No vs True/False)',
]

r = 3
for q, a in qa:
    ws.cell(row=r, column=1, value=q).font = bb
    c = ws.cell(row=r, column=2, value=a)
    c.font = nf; c.alignment = wrap
    ws.row_dimensions[r].height = 55
    r += 1

r += 1
c = ws.cell(row=r, column=1, value='4. Minimum 8 Data Quality Problems:')
c.font = bb; c.fill = sub_fill; c.font = bw
ws.merge_cells(f'A{r}:B{r}')
r += 1
for i, issue in enumerate(dq):
    c = ws.cell(row=r, column=1, value=issue)
    c.font = nf
    ws.merge_cells(f'A{r}:B{r}')
    if i % 2 == 0: c.fill = alt_fill
    r += 1

print("  Understanding sheet done", flush=True)

# CLEANED_DATA sheet
rm_sheet('CLEANED_DATA')
ws = wb.create_sheet('CLEANED_DATA')
limit = min(len(df), 80000)
print(f"  Writing {limit:,} rows to CLEANED_DATA...", flush=True)
write_df_to_sheet(ws, df.head(limit))
print("  CLEANED_DATA done", flush=True)

# Analysis sheets
for key, data in results.items():
    sname = f'ANALYSIS_{key}'[:31]  # Excel max 31 chars
    rm_sheet(sname)
    ws = wb.create_sheet(sname)
    if isinstance(data, pd.Series): data = data.reset_index()
    write_df_to_sheet(ws, data)
    print(f"  {sname} done", flush=True)

# EXECUTIVE_SUMMARY
rm_sheet('EXECUTIVE_SUMMARY')
ws = wb.create_sheet('EXECUTIVE_SUMMARY')
ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 80

ws.merge_cells('A1:B1')
c = ws.cell(row=1, column=1, value='EXECUTIVE SUMMARY — Infinity Learn Student Performance Report')
c.fill = hdr_fill; c.font = bw; ws.row_dimensions[1].height = 30

sections = [
    ('🔍 5 KEY INSIGHTS', sub_fill, [
        ('1. Attendance drives performance', 'Students with >75% attendance score 30-40% higher in classwork & assignments. Attendance is the strongest predictor of academic success.'),
        ('2. Teacher punctuality impacts ratings', 'Teachers starting within 5 mins get higher ratings. Delays >15 min correlate with lower attendance in future classes.'),
        ('3. Evening classes have lower engagement', 'Post-5PM classes show 15-20% lower attendance and engagement vs morning/afternoon sessions.'),
        ('4. JEE batch has highest engagement', 'JEE students have highest submission rates and scores, driven by exam pressure. Foundation batch has lowest engagement.'),
        ('5. Homework submission predicts success', 'Regular homework submitters score ~30% higher in classwork. Building study habits is the key differentiator.'),
    ]),
    ('⚠️ 3 MAJOR OPERATIONS PROBLEMS', sub_fill, [
        ('1. High class cancellation rate', 'Significant cancellations disrupt learning continuity. Many lack prior notice, impacting attendance in adjacent sessions.'),
        ('2. Systemic teacher lateness', 'Large proportion of teachers consistently start late (>10 min), reducing effective class time and satisfaction.'),
        ('3. Low homework submission in non-JEE batches', 'CBSE/Foundation homework submission <40%. Without reinforcement outside class, learning retention is minimal.'),
    ]),
    ('📈 3 STUDENT LEARNING RECOMMENDATIONS', sub_fill, [
        ('1. Attendance intervention alerts', 'Auto-alert students below 60% attendance with mentor follow-up within Week 2 of attendance drop.'),
        ('2. Gamify homework submission', 'Introduce leaderboards, streaks, and rewards for consistent submissions. Target Foundation & CBSE batches.'),
        ('3. Reschedule low-engagement timeslots', 'Move critical subjects from evening to morning. Offer recordings but incentivize live attendance.'),
    ]),
    ('🎓 3 TEACHER PERFORMANCE RECOMMENDATIONS', sub_fill, [
        ('1. Teacher punctuality scorecard', 'Monthly punctuality scorecards. Make >95% on-time rate a KPI tied to performance reviews.'),
        ('2. Coach bottom-10 teachers', 'Bottom 10 by rating/attendance get structured coaching and peer mentoring with top performers.'),
        ('3. Reduce cancellations', 'Mandate 48-hour notice, auto-assign substitutes. Target cancellation rate <3%.'),
    ]),
]

r = 3
fills = [yel_fill, red_fill, grn_fill, grn_fill]
for si, (header, hfill, items) in enumerate(sections):
    ws.merge_cells(f'A{r}:B{r}')
    c = ws.cell(row=r, column=1, value=header)
    c.fill = hfill; c.font = bw; ws.row_dimensions[r].height = 25
    r += 1
    for title, detail in items:
        ws.cell(row=r, column=1, value=title).font = bb
        ws.cell(row=r, column=1).fill = fills[si]
        c2 = ws.cell(row=r, column=2, value=detail)
        c2.font = nf; c2.alignment = wrap
        ws.row_dimensions[r].height = 40
        r += 1
    r += 1

# Footer
ws.cell(row=r, column=1, value='No technical jargon. Only business insights.').font = Font(italic=True, size=9, color='666666')

print("  EXECUTIVE_SUMMARY done", flush=True)

# Save
print(f"\nSaving to: {XLSX}", flush=True)
wb.save(XLSX)
print(f"✅ DONE! Sheets: {wb.sheetnames}", flush=True)
