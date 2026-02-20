#!/usr/bin/env python3
"""Fix pass: add score-based analyses that were missed"""
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import warnings
warnings.filterwarnings('ignore')

XLSX = "/home/wasim/Documents/github/DATA-INTERN-ASSIGNMENT/Assignment_data_dictionary.xlsx"

print("Reading...", flush=True)
df = pd.read_excel(XLSX, sheet_name='Raw_data')
print(f"Shape: {df.shape}", flush=True)

# Parse key columns
df['CLASS_START_DATETIME'] = pd.to_datetime(df['CLASS_START_DATETIME'], errors='coerce')
df['ACTUAL_START_DATETIME'] = pd.to_datetime(df['ACTUAL_START_DATETIME'], errors='coerce')
df['IS_PRESENT'] = df['CLASS_ATTENDANCE'].astype(str).str.upper().str.strip().isin(['PRESENT','1','Y','YES']).astype(int)

# Score columns
cw_score = 'CW_STUDENT_SCORE_ACHIEVED'
cw_max = 'CW_MAX_ACHIEVABLE_SCORE'
hw_score = 'STUDENT_SCORE_ACHIEVED'
hw_max = 'MAX_ACHIEVABLE_SCORE'
rating = 'PLEASE_RATE_YOUR_OVERALL_EXPERIENCE'
tutor = 'DID_THE_TUTOR_HELP_YOU_UNDERSTAND_THE_TOPIC_OF_THE_CLASS'
delay = 'TEACHER_IS_LATE_BY_MINS'

# Clean scores
for c in [cw_score, hw_score]:
    df[c] = pd.to_numeric(df[c], errors='coerce')
    df.loc[df[c] < 0, c] = 0

for s, m in [(cw_score, cw_max), (hw_score, hw_max)]:
    df[m] = pd.to_numeric(df[m], errors='coerce')
    mask = df[s] > df[m]
    df.loc[mask, s] = df.loc[mask, m]

# Score percentages
df['CW_SCORE_PCT'] = (df[cw_score] / df[cw_max] * 100).round(1)
df['HW_SCORE_PCT'] = (df[hw_score] / df[hw_max] * 100).round(1)

# Teacher punctuality
df[delay] = pd.to_numeric(df[delay], errors='coerce')
df['TEACHER_PUNCTUALITY'] = pd.cut(df[delay], bins=[-9999, 5, 15, 99999], labels=['On Time','Late','Very Late'])

# Engagement score = 30*(attendance) + 20*(CW submit rate) + 20*(HW submit rate) + 15*(avg score %) + 15*(rating/5)
df['CW_SUBMIT_RATE'] = (pd.to_numeric(df['NO_OF_CW_ASSIGNMENTS_SUBMITTED'], errors='coerce').fillna(0) /
                         pd.to_numeric(df['NUM_OF_CW_ASSIGNMENTS_GIVEN'], errors='coerce').replace(0, np.nan)).fillna(0).clip(0,1)
df['HW_SUBMIT_RATE'] = (pd.to_numeric(df['NO_OF_ASSIGNMENTS_SUBMITTED'], errors='coerce').fillna(0) /
                         pd.to_numeric(df['NUM_OF_ASSIGNMENTS_GIVEN'], errors='coerce').replace(0, np.nan)).fillna(0).clip(0,1)
df[rating] = pd.to_numeric(df[rating], errors='coerce')

df['ENGAGEMENT_SCORE'] = (
    30 * df['IS_PRESENT'] +
    20 * df['CW_SUBMIT_RATE'] +
    20 * df['HW_SUBMIT_RATE'] +
    15 * (df['CW_SCORE_PCT'].fillna(0) / 100) +
    15 * (df[rating].fillna(0) / 5)
).round(1)

print("Transformations done", flush=True)

# === ANALYSIS ===
results = {}

# 4.1 Student Behaviour (with scores!)
print("4.1 Student behaviour...", flush=True)
ss = df.groupby('STUDENT ID').agg(
    Total_Classes=('IS_PRESENT', 'count'),
    Attended=('IS_PRESENT', 'sum'),
    Att_Rate=('IS_PRESENT', 'mean'),
    Avg_CW_Score_Pct=('CW_SCORE_PCT', 'mean'),
    Avg_HW_Score_Pct=('HW_SCORE_PCT', 'mean'),
    CW_Submit_Rate=('CW_SUBMIT_RATE', 'mean'),
    HW_Submit_Rate=('HW_SUBMIT_RATE', 'mean'),
    Avg_Engagement=('ENGAGEMENT_SCORE', 'mean'),
).round(2)
ss['Att_Rate'] = (ss['Att_Rate'] * 100).round(1)
ss['CW_Submit_Rate'] = (ss['CW_Submit_Rate'] * 100).round(1)
ss['HW_Submit_Rate'] = (ss['HW_Submit_Rate'] * 100).round(1)

# Attendance vs score correlation
att_buckets = pd.cut(ss['Att_Rate'], bins=[0,25,50,75,100], labels=['0-25%','25-50%','50-75%','75-100%'])
att_score = ss.groupby(att_buckets).agg(
    Avg_CW_Score=('Avg_CW_Score_Pct', 'mean'),
    Avg_HW_Score=('Avg_HW_Score_Pct', 'mean'),
    Avg_CW_Submit=('CW_Submit_Rate', 'mean'),
    Avg_HW_Submit=('HW_Submit_Rate', 'mean'),
    Student_Count=('Total_Classes', 'count')
).round(1)
results['AttendVsScores'] = att_score.reset_index()
print(f"  Attend vs Scores:\n{att_score}", flush=True)

# Top 10% students
top10pct = ss.nlargest(int(len(ss)*0.1), 'Att_Rate')
top_avg = top10pct.mean().round(1)
bot_avg = ss.nsmallest(int(len(ss)*0.1), 'Att_Rate').mean().round(1)
comparison = pd.DataFrame({'Top10%': top_avg, 'Bottom10%': bot_avg, 'Difference': (top_avg - bot_avg).round(1)})
results['Top10vsBottom10'] = comparison.reset_index()
print(f"  Top 10% vs Bottom 10%:\n{comparison}", flush=True)

results['StudentBehaviour'] = ss.reset_index().head(500)
results['Top10pctStudents'] = top10pct.reset_index().head(100)

# 4.2 Teacher Performance (with scores + ratings!)
print("4.2 Teacher performance...", flush=True)
ts = df.groupby('TEACHER_NAME').agg(
    Classes_Taken=('IS_PRESENT', 'count'),
    Att_Rate=('IS_PRESENT', 'mean'),
    Avg_Rating=('PLEASE_RATE_YOUR_OVERALL_EXPERIENCE', 'mean'),
    Avg_Tutor_Rating=('DID_THE_TUTOR_HELP_YOU_UNDERSTAND_THE_TOPIC_OF_THE_CLASS', 'mean'),
    Avg_Delay_Mins=('TEACHER_IS_LATE_BY_MINS', 'mean'),
    Avg_CW_Score=('CW_SCORE_PCT', 'mean'),
    Avg_HW_Score=('HW_SCORE_PCT', 'mean'),
).round(2)
ts['Att_Rate'] = (ts['Att_Rate'] * 100).round(1)

# Punctuality rate
punct = df.groupby('TEACHER_NAME')['TEACHER_PUNCTUALITY'].apply(lambda x: (x == 'On Time').sum() / len(x) * 100).round(1)
ts['OnTime_Pct'] = punct

# Composite teacher score
ts['Teacher_Score'] = (
    0.3 * ts['Att_Rate'] / ts['Att_Rate'].max() * 100 +
    0.25 * ts['Avg_Rating'].fillna(0) / 5 * 100 +
    0.2 * ts['OnTime_Pct'] / 100 * 100 +
    0.15 * ts['Avg_CW_Score'].fillna(0) / ts['Avg_CW_Score'].max() * 100 +
    0.1 * ts['Avg_HW_Score'].fillna(0) / ts['Avg_HW_Score'].max() * 100
).round(1)

ts_r = ts.reset_index().sort_values('Teacher_Score', ascending=False)
results['TeacherPerformance'] = ts_r
results['Top10Teachers'] = ts_r.head(10)
results['Bottom10Teachers'] = ts_r.tail(10)

# 4.3 Exam Insights
print("4.3 Exam insights...", flush=True)
ei = df.groupby('EXAM').agg(
    Total_Sessions=('IS_PRESENT', 'count'),
    Att_Rate=('IS_PRESENT', 'mean'),
    Avg_CW_Score=('CW_SCORE_PCT', 'mean'),
    Avg_HW_Score=('HW_SCORE_PCT', 'mean'),
    CW_Submit_Rate=('CW_SUBMIT_RATE', 'mean'),
    HW_Submit_Rate=('HW_SUBMIT_RATE', 'mean'),
    Avg_Rating=('PLEASE_RATE_YOUR_OVERALL_EXPERIENCE', 'mean'),
    Avg_Engagement=('ENGAGEMENT_SCORE', 'mean'),
).round(2)
ei['Att_Rate'] = (ei['Att_Rate'] * 100).round(1)
ei['CW_Submit_Rate'] = (ei['CW_Submit_Rate'] * 100).round(1)
ei['HW_Submit_Rate'] = (ei['HW_Submit_Rate'] * 100).round(1)
results['ExamInsights'] = ei.reset_index()
print(f"  Exam insights:\n{ei}", flush=True)

# 4.4 Time-Based
print("4.4 Time-based...", flush=True)
df['HOUR'] = df['CLASS_START_DATETIME'].dt.hour
df['HOUR_BUCKET'] = df['HOUR'].apply(lambda h: 'Morning' if h < 12 else ('Afternoon' if h < 17 else 'Evening'))
df['DAY'] = df['CLASS_START_DATETIME'].dt.day_name()

time_att = df.groupby('HOUR_BUCKET').agg(
    Att_Rate=('IS_PRESENT','mean'),
    Avg_Rating=('PLEASE_RATE_YOUR_OVERALL_EXPERIENCE','mean'),
    Avg_Engagement=('ENGAGEMENT_SCORE','mean'),
).round(2)
time_att['Att_Rate'] = (time_att['Att_Rate']*100).round(1)
results['TimeAttendance'] = time_att.reset_index()
print(f"  Time attendance:\n{time_att}", flush=True)

day_att = df.groupby('DAY').agg(
    Att_Rate=('IS_PRESENT','mean'),
    Avg_Engagement=('ENGAGEMENT_SCORE','mean'),
).round(2)
day_att['Att_Rate'] = (day_att['Att_Rate']*100).round(1)
results['DayAttendance'] = day_att.reset_index()

# Punctuality vs Rating
punct_rat = df.groupby('TEACHER_PUNCTUALITY').agg(
    Avg_Rating=('PLEASE_RATE_YOUR_OVERALL_EXPERIENCE','mean'),
    Att_Rate=('IS_PRESENT','mean'),
    Count=('IS_PRESENT','count'),
).round(2)
punct_rat['Att_Rate'] = (punct_rat['Att_Rate']*100).round(1)
results['PunctualityVsRating'] = punct_rat.reset_index()
print(f"  Punctuality vs Rating:\n{punct_rat}", flush=True)

# 4.5 Other Metrics
print("4.5 Other metrics...", flush=True)
# Metric 1: Grade-wise attendance & performance
grade_stats = df.groupby('GRADE').agg(
    Att_Rate=('IS_PRESENT', 'mean'),
    Avg_CW_Score=('CW_SCORE_PCT', 'mean'),
    Avg_Engagement=('ENGAGEMENT_SCORE', 'mean'),
).round(2)
grade_stats['Att_Rate'] = (grade_stats['Att_Rate']*100).round(1)
results['GradeWiseStats'] = grade_stats.reset_index()

# Metric 2: Assignment completion vs score correlation
df['TOTAL_SUBMIT_RATE'] = ((df['CW_SUBMIT_RATE'] + df['HW_SUBMIT_RATE']) / 2 * 100).round(0)
submit_buckets = pd.cut(df['TOTAL_SUBMIT_RATE'], bins=[0, 25, 50, 75, 100], labels=['0-25%','25-50%','50-75%','75-100%'])
submit_score = df.groupby(submit_buckets).agg(
    Avg_CW_Score=('CW_SCORE_PCT','mean'),
    Avg_HW_Score=('HW_SCORE_PCT','mean'),
    Att_Rate=('IS_PRESENT','mean'),
    Count=('IS_PRESENT','count')
).round(1)
submit_score['Att_Rate'] = (submit_score['Att_Rate']*100).round(1)
results['SubmitVsScore'] = submit_score.reset_index()
print(f"  Submit vs Score:\n{submit_score}", flush=True)

# Metric 3: Week-of-month trends
df['WEEK'] = df['CLASS_START_DATETIME'].dt.day.apply(lambda d: f'W{min((d-1)//7+1,4)}' if pd.notna(d) else None)
week_stats = df.groupby('WEEK').agg(
    Att_Rate=('IS_PRESENT','mean'),
    Avg_Engagement=('ENGAGEMENT_SCORE','mean')
).round(2)
week_stats['Att_Rate'] = (week_stats['Att_Rate']*100).round(1)
results['WeekTrends'] = week_stats.reset_index()

print("Analysis done.", flush=True)

# === WRITE TO EXCEL ===
print("Writing to Excel...", flush=True)
wb = load_workbook(XLSX)

hdr = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
bw = Font(bold=True, color='FFFFFF', size=11)

def rm(name):
    if name in wb.sheetnames: del wb[name]

def write_df(ws, dataframe, start=1):
    for ci, col in enumerate(dataframe.columns, 1):
        c = ws.cell(row=start, column=ci, value=str(col))
        c.fill = hdr; c.font = bw
    for ri, row in enumerate(dataframe.itertuples(index=False), start+1):
        for ci, val in enumerate(row, 1):
            try:
                if pd.isna(val): val = None
                elif isinstance(val, (np.integer,)): val = int(val)
                elif isinstance(val, (np.floating,)): val = float(val)
            except: pass
            try: ws.cell(row=ri, column=ci, value=val)
            except: ws.cell(row=ri, column=ci, value=str(val))

for key, data in results.items():
    sname = f'ANALYSIS_{key}'[:31]
    rm(sname)
    ws = wb.create_sheet(sname)
    if isinstance(data, pd.Series): data = data.reset_index()
    write_df(ws, data)
    print(f"  {sname} created", flush=True)

print("Saving...", flush=True)
wb.save(XLSX)
print(f"✅ FIX DONE! All sheets: {wb.sheetnames}", flush=True)
